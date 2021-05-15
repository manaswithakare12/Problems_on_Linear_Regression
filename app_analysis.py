""" 
          ^    ^    ^    ^    ^    ^    ^    ^    ^    ^    ^    ^    ^    ^
         /B\  /e\  /t\  /t\  /e\  /r\  /T\  /e\  /a\  /c\  /h\  /i\  /n\  /g\
        <___><___><___><___><___><___><___><___><___><___><___><___><___><___>

Developed by: Ashish Kumar
"""
# %% Import dependencies
from src import *
import json
import pandas as pd
from openpyxl import load_workbook, drawing
import librosa
import numpy as np
import sys
from pydub import AudioSegment, silence

# %% method


def LoadTranscript():
    """Load transcription"""
    text = json.load(open(path + "/out/transcription.json", "rb"))
    emotion = pd.read_excel(
        path + "/out/EmotionDetection.xlsx", sheet_name="Emotion Detection"
    )
    return (text, emotion.iloc[0, 2])


def txtanalysis(text):
    """Text analysis"""
    AverageLengthOfSen = AverageLengthOfSentence(text)
    Wordrep, Wordreppercentage, length = WordRepetition(text, threshold=1)
    FillerWordPercentage, FillerWd, LengthFillerWord = FillerWord(text, length)
    Qnum = NumberOfQuestionsAsked(text)
    Adj, Speaking_rate = ChangeInPace(text)
    return (
        AverageLengthOfSen,
        Wordrep,
        Wordreppercentage,
        FillerWordPercentage,
        FillerWd,
        LengthFillerWord,
        Qnum,
        Adj,
        Speaking_rate,
    )


def PointMaker(
    AverageLengthOfSen,
    Speaking_rate,
    Qnum,
    Wordreppercentage,
    Wordrep,
    FillerWordPercentage,
    FillerWd,
    SilenceSpeech,
    Adj,
    mono,
    volume,
    emotion,
    eyecontact,
    gesture,
):
    """To get the point based on the meetrics value"""

    def PointAverageLengthOfSentence(AverageLengthOfSen):
        if AverageLengthOfSen < 8:
            point = 5
            insight = f"Average length of sentence is {AverageLengthOfSen} words"
            CoachingAdvice = f"Your sentence are too short, while its good to have short sentence in middle, but overall increase the number of words per sentence."
        elif AverageLengthOfSen >= 8 and AverageLengthOfSen < 14:
            point = 9
            insight = f"Average length of sentence is {AverageLengthOfSen} words"
            CoachingAdvice = "Try to include more words"
        elif AverageLengthOfSen >= 14 and AverageLengthOfSen < 25:
            point = 10
            insight = f"Average length of sentence is {AverageLengthOfSen} words"
            CoachingAdvice = "Your sentence length is perfect"
        elif AverageLengthOfSen >= 25 and AverageLengthOfSen < 40:
            point = 7
            insight = f"Average length of sentence is {AverageLengthOfSen} words"
            CoachingAdvice = "Reduce the number of words per sentence"
        else:
            point = 5
            insight = f"Average length of sentence is {AverageLengthOfSen} words"
            CoachingAdvice = "The average word per sentence is too much, please reduce them"
        return (point, insight, CoachingAdvice)

    def PointWordPerMinute(Speaking_rate):
        if Speaking_rate <= 80:
            point = 3
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "You have to improve almost twice. Start with tongue twisters, Breath deeply, Control your breath, breath less while you read or talk."
        elif Speaking_rate >= 80 and Speaking_rate < 100:
            point = 5
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "Take a long breath, clear your mind and read the script that you are going to speak in front of mirror."
        elif Speaking_rate >= 100 and Speaking_rate < 120:
            point = 6
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "You are almost there, just find the rhythm."
        elif Speaking_rate >= 120 and Speaking_rate < 130:
            point = 8
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "Just phrase carefully, also try with tongue twisters"
        elif Speaking_rate >= 130 and Speaking_rate < 155:
            point = 10
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "Keep it up! No recommendation."
        elif Speaking_rate >= 155 and Speaking_rate < 170:
            point = 8
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "For radio hosts, this is good, but for normal talk/teach, you might want to slow down a bit."
        elif Speaking_rate >= 170 and Speaking_rate < 180:
            point = 8
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "You are matching with Ted Speakers. If you are teaching or communicating in day to day life then you might want to slow down."
        else:
            point = 6
            insight = f"You spoke at an average of {Speaking_rate} words per minute."
            CoachingAdvice = "High speaking rate signifies that you are nervous, try to calm yourself and then speak."
        return (point, insight, CoachingAdvice)

    def PointQuestionAsked(Qnum):
        if Qnum == 0:
            point = 5
            insight = f"You asked {Qnum} questions."
            CoachingAdvice = "Please ask questions that will keep audience engaged."
        elif Qnum >= 1 and Qnum < 3:
            point = 10
            insight = f"You asked {Qnum} questions."
            CoachingAdvice = "Your audience must be engaged. Keep it up!"
        elif Qnum >= 3 and Qnum < 6:
            point = 9
            insight = f"You asked {Qnum} questions."
            CoachingAdvice = "If you are teaching students then its okay, but in normal conversation or talk you might want to decrease this number."
        elif Qnum >= 6 and Qnum <= 10:
            point = 7
            insight = f"You asked {Qnum} questions."
            CoachingAdvice = "Try to ask less questions."
        else:
            point = 6
            insight = f"You asked {Qnum} questions."
            CoachingAdvice = "Ask less question in that way audience will be focus more on what you have to say rather than spending more time in thinking about the questions."
        return (point, insight, CoachingAdvice)

    def PointWordrep(Wordreppercentage, Wordrep):
        if Wordreppercentage >= 1 and Wordreppercentage < 3:
            point = 10
            insight = f"Word repeation rate is {Wordreppercentage}%. Words repeated are: {Wordrep}"
            CoachingAdvice = "No Advice!"
        elif Wordreppercentage >= 3 and Wordreppercentage < 6:
            point = 8
            insight = f"Word repeation rate is {Wordreppercentage}%. Words repeated are: {Wordrep}"
            CoachingAdvice = "Avoid words repeatation."
        elif Wordreppercentage >= 6 and Wordreppercentage < 10:
            point = 6
            insight = f"Word repeation rate is {Wordreppercentage}%. Words repeated are: {Wordrep}"
            CoachingAdvice = "Try to work on the synonym of the words that you repeat."
        else:
            point = 4
            insight = f"Word repeation rate is {Wordreppercentage}%. Words repeated are: {Wordrep}"
            CoachingAdvice = "You need to increase your vocab."
        return (point, insight, CoachingAdvice)

    def PointFiller(FillerWordPercentage,FillerWd):
        if FillerWordPercentage < 3:
            point = 10
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 1"
        elif FillerWordPercentage >= 3 and FillerWordPercentage < 6:
            point = 8
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 2"
        elif FillerWordPercentage >= 6 and FillerWordPercentage < 12:
            point = 6
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 3"
        elif FillerWordPercentage >= 12 and FillerWordPercentage < 17:
            point = 5
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 4"
        elif FillerWordPercentage >= 17 and FillerWordPercentage < 25:
            point = 4
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 5"
        elif FillerWordPercentage >= 25 and FillerWordPercentage < 40:
            point = 3
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 6"
        elif FillerWordPercentage >= 40 and FillerWordPercentage < 75:
            point = 1
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 7"
        else:
            point = 0
            insight = f"You used filler words like {FillerWd} {FillerWordPercentage}% of time."
            CoachingAdvice = "CA FW 8"
        return (point, insight, CoachingAdvice)

    def PointSilence(SilenceSpeech):
        if SilenceSpeech == 1:
            point = 10
            insight = "There were no abrupt silence while you were speaking."
            CoachingAdvice = "Good, Keep it up!"
        else:
            point = 6
            insight = "There were abrupt silence while you were speaking."
            CoachingAdvice = "You can avoid it by slowing down while speaking which will give you ample of time to process. Alternatively, you can use filler words."
        return (point, insight, CoachingAdvice)

    def PointChangeInPace(Adj):
        if Adj == "Slow":
            point = 5
            insight = f"Your pace was {Adj}"
            CoachingAdvice = "Try to vary your pace with the number of words you speak every minute."
        elif Adj == "Good":
            point = 10
            insight = f"Your pace was {Adj}"
            CoachingAdvice = "Your pace is good. Keep it up!"
        elif Adj == "Fast":
            point = 5
            insight = f"Your pace was {Adj}"
            CoachingAdvice = "Slow down your pace, you may want to practice by speaking slowly and then speeding your transcript."
        return (point, insight, CoachingAdvice)

    def PointMonotonous(mono):
        if mono == 1:
            point = 10
            insight = "Your tone was not monotonous."
            CoachingAdvice = "No Advice. Keep it up!"
        else:
            point = 7
            insight = "Your tone was monotonous"
            CoachingAdvice = "Try to vary your tone (high/low), speaking in same tone makes session boring."
        return (point, insight, CoachingAdvice)

    def PointVolume(volume):  # threshold need to customize
        if volume == 1:
            point = 10
            insight = "Volume was good. We can hear you properly."
            CoachingAdvice = "Keep it up!"
        else:
            point = 6
            insight = "You sounded very low. It was difficult to hear you."
            CoachingAdvice = "You have to bring enery while speaking and make sure that you are heard properly by everyone in the audience."
        return (point, insight, CoachingAdvice)

    def PointEmotionalDetection(emotion): #psychological research
        if emotion >= 0 and emotion < 25:
            point = 4
            insight = f"You were calm or smiling {emotion}% of the time."
            CoachingAdvice = "You need to control your emotion while speaking. Smile and if you cannot then at least try to look calm."
        elif emotion >= 25 and emotion < 50:
            point = 6
            insight = f"You were calm or smiling {emotion}% of the time."
            CoachingAdvice = "People are engaged towards smiling people as they feel positive. You have to smile while speaking, few occasional jokes can help."
        elif emotion >= 50 and emotion < 75:
            point = 8
            insight = f"You were calm or smiling {emotion}% of the time."
            CoachingAdvice = "People love your smile and positivity. You can smile more often!"
        else:
            point = 10
            insight = f"You were calm or smiling {emotion}% of the time."
            CoachingAdvice = "Keep it up!"
        return (point, insight, CoachingAdvice)

    def PointEyeContact(eyecontact):
        if eyecontact >= 65:
            point = 10
            insight = (
                f"You were maintaining eye contact {round(eyecontact,2)}% of the time"
            )
            CoachingAdvice = "Making eye contact while speaking is a good approach. Keep it up!"
        else:
            point = 5
            insight = (
                f"You were maintaining eye contact {round(eyecontact,2)}% of the time"
            )
            CoachingAdvice = "Try to make eye contact, best way would be to practice in front of mirror and look straight into your eyes"
        return (point, insight, CoachingAdvice)

    def PointGesture(gesture):
        if gesture > 5:
            point = gesture
            insight = "You made gesture while speaking"
            CoachingAdvice = "Gesture is important part of effective communication. Keep it up!"
        else:
            point = gesture
            insight = "You didn't had any gesture."
            CoachingAdvice = "Start with hand gesture while talking, that would make your whole body engage and committed."
        return (point, insight, CoachingAdvice)

    (
        ValueAverageLengthOfSentence,
        ValueWordPerMinute,
        ValueQuestionAsked,
        ValueWordrep,
        ValueFiller,
        ValueSilence,
        ValueChangeInPace,
        ValueMonotonous,
        ValueVolume,
        ValueEmotionalDetection,
        ValueEyeContact,
        ValueGesture,
    ) = (
        PointAverageLengthOfSentence(AverageLengthOfSen),
        PointWordPerMinute(Speaking_rate),
        PointQuestionAsked(Qnum),
        PointWordrep(Wordreppercentage,Wordrep),
        PointFiller(FillerWordPercentage,FillerWd),
        PointSilence(SilenceSpeech),
        PointChangeInPace(Adj),
        PointMonotonous(mono),
        PointVolume(volume),
        PointEmotionalDetection(emotion),
        PointEyeContact(eyecontact),
        PointGesture(gesture),
    )

    return (
        ValueAverageLengthOfSentence,
        ValueWordPerMinute,
        ValueQuestionAsked,
        ValueWordrep,
        ValueFiller,
        ValueSilence,
        ValueChangeInPace,
        ValueMonotonous,
        ValueVolume,
        ValueEmotionalDetection,
        ValueEyeContact,
        ValueGesture,
    )


def compute_power_db(x, Fs, win_len_sec=0.1, power_ref=10 ** (-12)):
    """Computation of the signal power in dB

    Args:
        x (np.ndarray): Signal (waveform) to be analyzed
        Fs (scalar): Sampling rate
        win_len_sec (float): Length (seconds) of the window (Default value = 0.1)
        power_ref (float): Reference power level (0 dB) (Default value = 10**(-12))

    Returns:
        power_db (np.ndarray): Signal power in dB
    """
    win_len = round(win_len_sec * Fs)
    win = np.ones(win_len) / win_len
    power_db = 10 * np.log10(np.convolve(x ** 2, win, mode="same") / power_ref)

    power_db = power_db[np.isfinite(power_db)]

    return np.mean(power_db[power_db > 0])


def EyeCon(threshold=95):
    """To canculate Eye contact percentage"""
    frameFiles = dirFile(path + "/db/out/", "json")
    EyeContactJSON = np.array(
        [
            json.load(open(f, "r"))["FaceDetails"][0]["EyesOpen"]["Confidence"]
            for f in frameFiles
        ]
    )
    PercentEyeContact = (
        len(EyeContactJSON[EyeContactJSON > threshold]) * 100 / len(EyeContactJSON)
    )
    return PercentEyeContact


def PitchVolume():
    """Find pitch and volume"""

    y, sr = librosa.load(path + "/audio/audioMono.wav")
    pitches, _ = librosa.core.piptrack(y=y, sr=sr, fmin=75, fmax=1600)
    np.set_printoptions(threshold=sys.maxsize)
    p = pitches[np.nonzero(pitches)]

    maxP = np.max(p)
    minP = np.min(p)
    stdP = np.std(p)

    if (maxP - minP) >= 3 * stdP:
        monotonous = 0
    else:
        monotonous = 1

    Loudness = compute_power_db(y, sr)

    # https://www.maxyourhometime.com/choosing-loudest-speaker-sensitivity-and-power-wattage/#:~:text=SPL%20is%20expressed%20in%20terms,%2Dreflective%2C%20soundproof%20room).

    if Loudness >= 75 and Loudness < 100:
        volume = 1
    else:
        volume = 0

    return (monotonous, volume)


def silenceDetection():
    """To find out adrupt silence"""
    from pydub import AudioSegment, silence

    sound = AudioSegment.from_wav(path + "/audio/audioMono.wav")

    chunks = silence.split_on_silence(
        sound, min_silence_len=4 * 1000, silence_thresh=-16
    )  # Change the parameter as per your need

    return len(chunks)


def PopulateExcel(filename="Sample"):
    """Populate the analysis result in excel"""
    text, emotion = LoadTranscript()
    (
        AverageLengthOfSentence,
        Wordrep,
        Wordreppercentage,
        FillerWordPercentage,
        FillerWd,
        LengthFillerWord,
        Qnum,
        Adj,
        Speaking_rate,
    ) = txtanalysis(text)

    wb = load_workbook(path + "/templates/Report.xlsx")
    ws = wb.active
    # Assigning data directly to cell

    Name = str(
        input("Enter the name of speaker [default: Prashant Kumar Dey]: ")
        or "Prashant Kumar Dey"
    )
    ws["A4"] = Name

    # IMG_FILE_ADDED = str(
    #     input(
    #         "Add image snippet path [e.g. C:/Users/imash/Documents/speech2text/templates/frame_1.4_7.jpg]: "
    #     )
    # )

    IMG_FILE_ADDED = dirFile(path + "/db/input/videos/OutputDump/", ".jpg")[0].replace(
        "\\", "/"
    )

    try:
        img = drawing.image.Image(IMG_FILE_ADDED)
        img.height = 50
        img.width = 50
        img.anchor = "B4"
        ws.add_image(img)
    except:
        pass

    ws["B4"] = IMG_FILE_ADDED

    SilenceSpeechDummy = int(
        input(
            "Enter the point in between 1 to 10 based on adrupt silence in Audio[default: 0, If you want app to do it for you then type 99]: "
        )
        or "0"
    )  # Need to calculate

    gestureDummy = int(
        input(
            "Enter the point in between 1 to 10 based on gesture done by Speaker[default: 0]: "
        )
        or "0"
    )  # Need to calculate

    print("[info..] Analysis started !!")

    (
        monotonous,
        volume,
    ) = PitchVolume()  # Function to collect information on Pitch and Volume

    mono = monotonous
    volume = volume
    eyecontact = EyeCon(threshold=80)
    gesture = gestureDummy  # Need to calculate

    if SilenceSpeechDummy != 99:
        SilenceSpeech = 1

    # To find out silence is there or not in audio
    if SilenceSpeechDummy == 99:
        chunkLength = silenceDetection()
        if chunkLength > 0:
            SilenceSpeech = 1
        else:
            SilenceSpeech = 0

    (
        ValueAverageLengthOfSentence,
        ValueWordPerMinute,
        ValueQuestionAsked,
        ValueWordrep,
        ValueFiller,
        ValueSilence,
        ValueChangeInPace,
        ValueMonotonous,
        ValueVolume,
        ValueEmotionalDetection,
        ValueEyeContact,
        ValueGesture,
    ) = PointMaker(
        AverageLengthOfSentence,
        Speaking_rate,
        Qnum,
        Wordreppercentage,
        Wordrep,
        FillerWordPercentage,
        FillerWd,
        SilenceSpeech,
        Adj,
        mono,
        volume,
        emotion,
        eyecontact,
        gesture,
    )

    if SilenceSpeechDummy == 99:
        SilenceSpeechDummy = ValueSilence[0]

    # overallScore = (
    #     ValueAverageLengthOfSentence[0]
    #     + ValueWordPerMinute[0]
    #     + ValueQuestionAsked[0]
    #     + ValueWordrep[0]
    #     + ValueFiller[0]
    #     + ValueSilence[0]
    #     + ValueChangeInPace[0]
    #     + ValueMonotonous[0]
    #     + ValueVolume[0]
    #     + ValueEmotionalDetection[0]
    #     + ValueEyeContact[0]
    #     + ValueGesture[0]
    # ) / 120  # Overall score

    # 0 index for point, 1 for Insight and 2 for Advice
    overallScore = (
        ValueAverageLengthOfSentence[0]
        + ValueWordPerMinute[0]
        + ValueQuestionAsked[0]
        + ValueWordrep[0]
        + ValueFiller[0]
        + SilenceSpeechDummy
        + ValueChangeInPace[0]
        + ValueMonotonous[0]
        + ValueVolume[0]
        + ValueEmotionalDetection[0]
        + ValueEyeContact[0]
        + gestureDummy
    ) / 120  # Overall score

    ws["C4"] = round(overallScore, 2)

    # Advice
    ws["H4"] = ValueWordPerMinute[2] + ". " + ValueChangeInPace[2]  # Pace
    ws["K4"] = ValueFiller[2] + ". " + ValueSilence[2]  # Confidence
    ws["M4"] = ValueQuestionAsked[2]  # Interactivity
    ws["P4"] = ValueMonotonous[2] + ". " + ValueVolume[2]  # Energy
    ws["S4"] = ValueWordrep[2] + ". " + ValueAverageLengthOfSentence[2]  # Conciseness
    ws["U4"] = ValueEmotionalDetection[2]  # Emotion
    ws["X4"] = ValueGesture[2] + ". " + ValueEyeContact[2]  # Gesture

    # Insight
    ws["G4"] = ValueChangeInPace[1]
    ws["I4"] = ValueSilence[1]
    ws["N4"] = ValueMonotonous[1]
    ws["O4"] = ValueVolume[1]
    ws["V4"] = ValueEyeContact[1]
    ws["W4"] = ValueGesture[1]

    # Analysis data
    ws["F4"] = f"{Adj} average pace of {Speaking_rate} words per min"
    ws[
        "J4"
    ] = f"You used {FillerWordPercentage}% of filler words. {len(FillerWd)} Filler words - {str(FillerWd).replace('[','').replace(']','')} were used for {LengthFillerWord} times"
    ws["L4"] = f"You asked {Qnum} questions"
    ws[
        "Q4"
    ] = f"Few words were repeated. {str(Wordrep).replace('[','').replace(']','')} were repeated most which was {Wordreppercentage}% of your speech"
    ws["R4"] = f"Your average sentence length was {AverageLengthOfSentence}"
    ws["T4"] = f"You were happy and calm {round(emotion,2)}% of the time"

    procontable = pd.DataFrame(
        [
            {
                "lengthofsentence": ValueAverageLengthOfSentence[0], #0 means point, 1 means insights & 2 means advice
                "wordperminute": ValueWordPerMinute[0],
                "questionasked": ValueQuestionAsked[0],
                "wordrep": ValueWordrep[0],
                "filler": ValueFiller[0],
                "silence": ValueSilence[0],
                "changeinpace": ValueChangeInPace[0],
                "monotonous": ValueMonotonous[0],
                "volume": ValueVolume[0],
                "emotiondetection": ValueEmotionalDetection[0],
                "eyecontact": ValueEyeContact[0],
                "gesture": ValueGesture[0],
            },
            {
                "lengthofsentence": ValueAverageLengthOfSentence[1],
                "wordperminute": ValueWordPerMinute[1],
                "questionasked": ValueQuestionAsked[1],
                "wordrep": ValueWordrep[1],
                "filler": ValueFiller[1],
                "silence": ValueSilence[1],
                "changeinpace": ValueChangeInPace[1],
                "monotonous": ValueMonotonous[1],
                "volume": ValueVolume[1],
                "emotiondetection": ValueEmotionalDetection[1],
                "eyecontact": ValueEyeContact[1],
                "gesture": ValueGesture[1],
            },
        ]
    ).T

    procontable.rename(columns={0: "points", 1: "advice"}, inplace=True)
    procontable.index.name = "metrics"
    procontable.reset_index(inplace=True)

    pro = (
        # procontable[procontable["points"] > 5]
        procontable.sort_values(by=["points"], ascending=False)["advice"].head(2)
    )
    con = (
        # procontable[procontable["points"] <= 5]
        procontable.sort_values(by=["points"], ascending=True)["advice"].head(2)
    )

    # Pros-Cons
    ws["D4"] = (
        str(list(pro))
        .replace("[", "👍 ")
        .replace("]", "")
        .replace(",", "\n👍 ")
        .replace(" '", "")
        .replace("'", "")
    )  # Pros
    ws["E4"] = (
        str(list(con))
        .replace("[", "👎 ")
        .replace("]", "")
        .replace(",", "\n👎 ")
        .replace(" '", "")
        .replace("'", "")
    )  # Cons

    wb.save(path + f"/out/{filename}.xlsx")
    print("Workbook created successfully")


if __name__ == "__main__":

    PopulateExcel()
